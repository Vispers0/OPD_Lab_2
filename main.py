import logging

from aiogram import Bot, Dispatcher, executor, types
from aiogram.types import InlineKeyboardMarkup, InlineKeyboardButton

import openpyxl
from openpyxl.workbook import Workbook
from openpyxl.styles import Font, Alignment

API_TOKEN = "6139629143:AAEcQmSzz53RmYsYAcv_CU_-avJQs6c6kWU"

logging.basicConfig(level=logging.INFO)

bot = Bot(token=API_TOKEN)
disp = Dispatcher(bot)

_sum = 0
_period = 0
_percent = 0

_initial_budget = 0

_callback_data = None


@disp.message_handler(commands=['start', 'menu'])
async def menu(message: types.Message):
    markup = InlineKeyboardMarkup(row_width=1)

    btn_credit = InlineKeyboardButton(text="Кредитный калькулятор", callback_data="credit")
    btn_deposit = InlineKeyboardButton(text="Калькулятор вкладов", callback_data="deposit")
    btn_game = InlineKeyboardButton(text="Игра \"52 недели богатства\"", callback_data="game")

    markup.add(btn_credit, btn_deposit, btn_game)
    await message.answer("Здравствуйте! Выберите необходимую программу:", reply_markup=markup)


@disp.callback_query_handler()
async def callback_handler(callback: types.CallbackQuery):
    global _callback_data
    _callback_data = None

    if callback.data == "credit":
        _callback_data = 0
        await bot.delete_message(callback.message.chat.id, callback.message.message_id)
        await callback.message.answer("Введите сумму кредита в формате \"/sum <Сумма>\"")
    if callback.data == "deposit":
        _callback_data = 1
        await bot.delete_message(callback.message.chat.id, callback.message.message_id)
        await callback.message.answer('Введите сумму вклада в формате \"/sum <Сумма>\"')
    if callback.data == "game":
        await bot.delete_message(callback.message.chat.id, callback.message.message_id)
        await callback.message.answer('*Правила игры*:\n'
                                      '1) Вы вводите сумму, с которой хотите начать накопления, используя команду /initbud <Сумма>\n'
                                      '❗️*ВНИМАНИЕ*❗ с каждым новым днём сумма, которую необходимо будет отложить, будет увеличиваться на начальную сумму. Выбирайте начальный бюджет с умом\n'
                                      '2) Сумма накоплений расписывается на 52 недели (1 год) вперёд\n'
                                      '3) Бот отправляет вам готовую таблицу накоплений\n'
                                      '4) Начиная с понедельника следующей недели, вы откладываете каждый день указанную сумму\n'
                                      'Если вы ознакомились с правилами и готовы, введите начальный бюджет\n'
                                      'Если вы не хотите продолжать, то введите команду /menu, чтобы вернуться в главное меню', parse_mode="Markdown")


@disp.message_handler(commands="sum")
async def get_sum(message: types.Message):
    global _sum
    _sum = 0
    args = message.text.split()
    _sum = float(args[1])
    await message.answer("Введите срок кредита/вклада в годах в формате \"/period <Срок>\"")


@disp.message_handler(commands="period")
async def get_period(message: types.Message):
    global _period
    _period = 0
    args = message.text.split()
    _period = float(args[1])
    await message.answer("Введите процентную ставку кредита/вклада (годовые) в формате \"/percent <Процентная ставка>\"")


@disp.message_handler(commands="percent")
async def get_percent(message: types.Message):
    global _percent
    global _callback_data
    _percent = 0
    args = message.text.split()
    _percent = float(args[1])

    match _callback_data:
        case 0:
            await credit(message)
        case 1:
            await deposit(message)


@disp.message_handler(commands="initbud")
async def rich_game(message: types.Message):
    global _initial_budget
    _initial_budget = 0

    args = message.text.split()
    _initial_budget = float(args[1])
    total = _initial_budget

    result_table = Workbook()
    active_list = result_table.active

    weekdays = ["№ Недели", "Понедельник", "Вторник", "Среда", "Четверг", "Пятница", "Суббота", "Воскресенье"]
    active_list.append(weekdays)

    text_style = Font(bold=True)
    alignment = Alignment(horizontal="center", vertical="center", shrink_to_fit=True)

    for i in range(2, 54):
        row = [i - 1]
        for j in range(0, 7):
            row.append(int(total))
            total += _initial_budget

        active_list.append(row)

    for i in range(1, 9):
        active_list.cell(row=1, column=i).font = text_style

    for i in range(1, 54):
        active_list.cell(row=i, column=1).font = text_style

    for i in range(1, 54):
        for j in range(1, 9):
            active_list.cell(row=i, column=j).alignment = alignment

    result_table.save("52_недели_богатства.xlsx")
    file = open("./52_недели_богатства.xlsx", "rb")

    await message.answer("Вот ваша таблица!")
    await bot.send_document(message.chat.id, file)
    await message.answer("Введите команду /menu, чтобы вернуться в главное меню")


async def credit(message: types.Message):
    await message.answer("Сумма кредита: " + str(_sum) + " руб.\n"
                         "Срок кредита: " + str(_period) + " лет\n"
                         "Процент кредита: " + str(_percent) + "%")
    monthly_pay = _sum * ((_percent / (100 * 12)) / (1 - pow((1 + _percent / (100 * 12)), (-_period * 12))))
    summary_pay = monthly_pay * _period * 12

    await message.answer("Ежемесячная выплата: " + str(round(monthly_pay, 3)) + " руб.\n"
                         "Общая сумма выплат: " + str(round(summary_pay, 3)) + " руб.\n"
                         "Переплата: " + str(round(summary_pay - _sum, 3)) + " руб.")
    await message.answer("Введите /menu, чтобы вернуться в главное меню")


async def deposit(message: types.Message):

    await message.answer("Сумма вклада: " + str(_sum) + " руб.\n"
                         "Срок вклада: " + str(_period) + " лет\n"
                         "Процент вклада: " + str(_percent) + "%")

    total = _sum
    income = 0

    for i in range(0, int(_period)):
        income = (total * _percent) / 100
        total += income

    await message.answer("Сумма вклада к окончанию срока: " + str(total) + " руб.\n"
                         "Доход: " + str(total - _sum) + " руб.")
    await message.answer("Введите /menu, чтобы вернуться в главное меню")


if __name__ == "__main__":
    executor.start_polling(disp, skip_updates=True)
